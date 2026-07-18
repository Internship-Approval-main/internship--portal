from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, model_validator
from typing import Optional, List, Literal
import pymongo
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
load_dotenv()

app = FastAPI()

# Enable CORS for frontend website communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB Connection Setup
MONGO_DB_NAME = "internship_db"
MONGO_COLLECTION_NAME = "internship_records"
MASTER_COLLECTION_NAME = "master_database"

client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client[MONGO_DB_NAME]
collection = db[MONGO_COLLECTION_NAME]
master_collection = db[MASTER_COLLECTION_NAME]

def get_missing_students():

    pipeline = [
        {"$match": {"semester": "8"}},
        {
            "$lookup": {
                "from": MONGO_COLLECTION_NAME,
                "localField": "srn",
                "foreignField": "srn",
                "as": "submission"
            }
        },
        {"$match": {"submission": {"$size": 0}}}
    ]

    return list(master_collection.aggregate(pipeline))
def send_email(student_email, student_name, srn):

    sender_email = os.getenv("EMAIL_USER")
    sender_password = os.getenv("EMAIL_PASS")

    message = MIMEMultipart()

    message["From"] = sender_email
    message["To"] = student_email
    message["Subject"] = "Internship Submission Reminder"

    body = f"""
Dear {student_name},

Our records show that you have not yet submitted your internship details.

Please log in to the Internship Portal and complete your submission as soon as possible.

SRN: {srn}

Regards,
Internship Management Team
"""

    message.attach(MIMEText(body, "plain"))

    try:

        server = smtplib.SMTP("smtp.gmail.com", 587)

        server.starttls()

        server.login(sender_email, sender_password)

        server.sendmail(
            sender_email,
            student_email,
            message.as_string()
        )

        server.quit()

        print(f"Email sent to {student_name}")
        return True

    except Exception as e:

        print(f"Failed to send email to {student_name}: {e}")
        return False
# ==========================================
# 1. THE SUB-BLUEPRINT (UPDATED VALIDATION)
# ==========================================
class InternshipDetails(BaseModel):
    campus_type: Literal["on campus", "off campus"]               
    internship_type: Literal["industry", "research"]           
    start_date: str                
    end_date: str                  
    
    # Globally required across both industry and research
    company: str                           # Company Name OR Institute Name
    role: str                              # Job Role or Position
    manager_name: str                      # Manager Name OR Supervisor Name
    manager_email: str                     # Manager Email OR Supervisor Email
    
    # Conditional fields
    company_link: Optional[str] = None     # Mandatory only for: Off Campus + Industry
    
    # Dropdown validation matching your exact menu UI options
    research_category: Optional[Literal[
        "CCBD", "ISFCR", "CSDML", "Pi Labs", "CHeal", "IOT", "PVL Labs", "RAAS", "Other"
    ]] = None 
    
    research_centre: Optional[str] = None   # Mandatory only if research_category is "Other"

    @model_validator(mode="after")
    def validate_conditional_fields(self):
        itype = self.internship_type.strip().lower()
        ctype = self.campus_type.strip().lower()

        # --- RULE 1: OFF-CAMPUS INDUSTRY ---
        if ctype == "off campus" and itype == "industry":
            if not self.company_link or not self.company_link.strip():
                raise ValueError(
                    "Company Website Link is mandatory for off-campus industry internships."
                )

        # --- RULE 2: RESEARCH ---
        elif itype == "research":
            if not self.research_category:
                raise ValueError("Please select a valid Research Centre from the menu.")
            
            if self.research_category == "Other":
                if not self.research_centre or not self.research_centre.strip():
                    raise ValueError(
                        "Please enter the custom Research Centre Name since you selected 'Other'."
                    )

        return self


# ==========================================
# 2. THE MAIN PROFILE SUBMISSION BLUEPRINT
# ==========================================
class StudentInternshipSubmission(BaseModel):
    srn: str                       
    student_name: str
    student_email: str
    semester: str                  
    placements: List[InternshipDetails] = Field(..., max_length=2) # Fixed the deprecation warning here!


# Helper function to build a formatted Excel workbook of missing students
def build_missing_students_excel(missing_students: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Submissions"

    headers = ["SRN", "Student Name", "Student Email", "Semester"]
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, student in enumerate(missing_students, start=2):
        ws.cell(row=row_idx, column=1, value=student.get("srn", "")).font = Font(name="Arial")
        ws.cell(row=row_idx, column=2, value=student.get("student_name", "")).font = Font(name="Arial")
        ws.cell(row=row_idx, column=3, value=student.get("student_email", "")).font = Font(name="Arial")
        ws.cell(row=row_idx, column=4, value=student.get("semester", "")).font = Font(name="Arial")

    widths = [16, 28, 32, 12]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ==========================================
# 3. ROUTE TO SAVE SUBMISSIONS TO MONGODB
# ==========================================
@app.post("/submit-record")
async def save_student_internships(submission: StudentInternshipSubmission):
    # Converts the frontend incoming payload data into a clean Python dictionary map
    doc = submission.model_dump()
    
    # Clean and normalize the strings before storing them in MongoDB
    for placement in doc["placements"]:
        placement["campus_type"] = placement["campus_type"].strip().lower()
        placement["internship_type"] = placement["internship_type"].strip().lower()
        
        # Strip string white spaces safely from our brand new conditional fields
        if placement.get("company_link"):
            placement["company_link"] = placement["company_link"].strip()
            
        if placement.get("research_centre"):
            placement["research_centre"] = placement["research_centre"].strip()
            
        # (Note: research_category is left in its exact uppercase/case format like "CCBD")

    # Dynamic Upsert: Finds student record by SRN. If found, it updates it. If missing, it creates it.
    result = collection.update_one(
        {"srn": doc["srn"]}, 
        {"$set": doc},        
        upsert=True           
    )
    
    if result.matched_count > 0:
        return {"status": "success", "message": f"Updated internship data records for {doc['student_name']}"}
    else:
        return {"status": "success", "message": f"Successfully registered system entry for {doc['student_name']}"}


# ==========================================
# 4. ROUTE FOR ADMIN EXPORTS
# ==========================================
@app.post("/admin/export-missing-students")
async def export_missing_students():
    missing_students = get_missing_students()

    if not missing_students:
        raise HTTPException(status_code=404, detail="All 8th semester students have submitted their details!")

    excel_buffer = build_missing_students_excel(missing_students)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"missing_internship_submissions_{timestamp}.xlsx"

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
@app.post("/admin/send-reminders")
async def send_reminders():

    missing_students = get_missing_students()

    if not missing_students:
        return {
            "status": "success",
            "emailsSent": 0,
            "message": "All students have already submitted their internship details."
        }

    sent_count = 0

    for student in missing_students:

        if send_email(
            student["student_email"],
            student["student_name"],
            student["srn"]
        ):
            sent_count += 1

    return {
        "status": "success",
        "emailsSent": sent_count,
        "message": "Reminder emails sent successfully."
    }